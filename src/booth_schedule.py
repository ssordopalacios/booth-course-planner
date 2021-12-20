import os
import numpy as np
import pandas as pd
import openpyxl
from . import helper


def main():
    """Read and clean the BoothSchedule file

    Returns:
        DataFrame: The cleaned BoothSchedule
    """

    # Read the schedule file
    fname = os.path.join("data", "BoothSchedule.xlsx")
    print(f"Reading Booth Schedule from {fname}")
    df = pd.read_excel(fname)

    # Break year and quarter
    df[["Quarter", "Year"]] = df["Quarter"].str.split(" ", expand=True)
    df["Year"] = df["Year"].astype(int)
    df["Quarter"] = df["Quarter"].map(helper.qtr_to_quarter)

    # Break course number and section and add program
    df[["Course", "Section"]] = df["Section"].str.split("-", expand=True)
    df["Course"] = df["Course"].astype(int)
    df["Section"] = df["Section"].astype(int)
    df["Program"] = df["Section"].map(helper.section_to_program)

    # Break instructor into last and first name
    df[["Last Name", "First Name"]] = df["Instructor"].str.split(
        ", ", expand=True
    )

    # Break apart day and time
    df[["Day", "Time"]] = df["Meeting Day/Time"].str.split(" ", expand=True)
    calendar_dict = {
        "M": "Monday",
        "T": "Tuesday",
        "W": "Wednesday",
        "TH": "Thursday",
        "F": "Friday",
        "S": "Saturday",
        "TTH": "Tuesday/Thursday",
        "MW": "Monday/Wednesday",
        "": "",
    }
    df["Day"].replace(calendar_dict, inplace=True)

    # TODO: Handle half week courses
    # TODO: Handle pre-reqs and negative pre-reqs
    # Get syllabi hyperlinks in column "L"
    df["Syllabus"] = hyperlinks(fname)

    # Remove unwanted columns
    df = df.drop(
        columns=["Unnamed: 5", "Unnamed: 6", "Meeting Day/Time", "Instructor"]
    )

    # Convert data types
    df = df.convert_dtypes()
    df = df.fillna(np.nan)

    # Clean up the title
    df["Title"] = df["Title"].map(helper.remove_ascii)
    df["Note"] = df["Note"].map(helper.remove_ascii)
    df["Prerequisites"] = df["Prerequisites"].map(helper.remove_ascii)

    # Order columns
    column_ordering = helper.column_ordering(df.columns)
    df = df[column_ordering]

    # Sort the columns
    column_sorting = helper.column_sorting(df.columns)
    df = df.sort_values(column_sorting)

    return df


def hyperlinks(fname):

    # Open the workbook
    wb = openpyxl.load_workbook(fname)
    ws = wb["Sheet1"]

    # Loop over each row and pull out the hyperlink
    links = []
    rows = len(ws["L"])
    for r in range(2, rows + 1):
        try:
            link = ws.cell(row=r, column=12).hyperlink.target
        except AttributeError:
            link = ""
        links.append(link)

    wb.close()
    return pd.Series(links)
